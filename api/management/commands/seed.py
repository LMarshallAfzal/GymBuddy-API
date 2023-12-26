import openpyxl
from django.core.management.base import BaseCommand
from tqdm import tqdm
from api.models import Exercise

class Command(BaseCommand):
    help = 'Seeds the database with exercise data from XLSX file'

    def handle(self, *args, **options):
        workbook = openpyxl.load_workbook('/home/leonard/Projects/GymBuddy-API/exercise_data.xlsx')
        sheet = workbook.active

        total_rows = sheet.max_row - 1
        with tqdm(total=total_rows, desc='Seeding Exercises') as pbar:
            for row in sheet.iter_rows(min_row=2):
                name = row[0].value
                description = row[1].value
                image_1 = row[2].value
                image_2 = row[3].value
                image_3 = row[4].value
                image_4 = row[5].value
                type = row[6].value
                muscle_group = row[7].value
                equipment = row[8].value
                level = row[9].value

                Exercise.objects.create(
                    name=name,
                    description=description,
                    type=type,
                    muscle_group=muscle_group,
                    equipment=equipment,
                    level=level,
                    image1=image_1,
                    image2=image_2,
                    image3=image_3,
                    image4=image_4
                )

                pbar.update(1)

        self.stdout.write(self.style.SUCCESS('Database seeded successfully!'))

    
    